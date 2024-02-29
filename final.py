from itertools import cycle
import requests
import json
import pymsteams
from datetime import datetime, timedelta
import urllib3
import openpyxl
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


'''
╔════════════════════════════════════════════════════════════════════════╗
║                    Aging Tickets Auto Assignment beta                  ║
╠════════════════════════════════════════════════════════════════════════╣
║                        last update: 15-Jan-2024                        ║
╠════════════════════════════════════════════════════════════════════════╣
║ This script pulls the unassigned aging tickets per shift.              ║
║ This information is then used to assign tickets to those on shift.     ║
╚════════════════════════════════════════════════════════════════════════╝
'''



DEBUG_MODE = False
current_time = datetime.now()
ap_on_shift = []
eu_on_shift = []
us_on_shift = []
final_assigned = {}
WEBHOOK_URL = "https://pgone.webhook.office.com/webhookb2/b324da08-4c77-4788-81de-57876db9d901@3596192b-fdf5-4e2c-a6fa-acb706c963d8/IncomingWebhook/935f0ea5492b43cd88f7c53660eab360/f462f2fe-b20b-4ccb-b306-1993fcfab318"



def check_shift():
    current_date = datetime.now().date()  
    workbook = openpyxl.load_workbook('schedule_final.xlsx')
    sheet = workbook['Sheet1']
    data_list = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        data_list.append(row)

    #compare dates
    counter = 0
    for i in data_list[0]:
        if i is None or isinstance(i,str):
            counter += 1
            continue
        elif i.date() == current_date:
            break
        else:
            counter += 1

    #counter index gives the current date
    ap_shift_sched = ['AP9', 'AP9*', 'AP4+4', 'MID']
    eu_shift_sched = ['EU9', 'EU9*', 'EU4+4']
    us_shift_sched = ['US1', 'US2', 'US3', 'US4', 'US5']
    for index, row in enumerate(data_list):
        if row[counter] in ap_shift_sched:
            cell = sheet.cell(row=index+1, column=counter+1)
            fill_color = cell.fill.start_color.rgb
            ap_on_shift.append([row[2]])
            # mnl_on_shift.append([row[2],row[3],row[counter],fill_color])

        elif row[counter] in us_shift_sched:
            cell = sheet.cell(row=index+1, column=counter+1)
            fill_color = cell.fill.start_color.rgb
            if fill_color == 'FF70AD47':
                # us_on_shift.append([row[2],row[3],row[counter],fill_color])
                us_on_shift.append([row[2]])

        elif row[counter] in eu_shift_sched:
            cell = sheet.cell(row=index+1, column=counter+1)
            fill_color = cell.fill.start_color.rgb
            eu_on_shift.append([row[2]])

def assign_ticket(engineer_on_shift, tickets):
    engineer_cycle = cycle(engineer_on_shift)
    for item in tickets:
        engineer = next(engineer_cycle)
        result_string = ', '.join(engineer)
        if result_string not in final_assigned:
            final_assigned[result_string] = [item]
        else:
            final_assigned[result_string].append(item)
    return final_assigned

def parse_tickets(tickets):
    data = []
    for item in tickets["result"]:
        data.extend([[item["number"], item["short_description"]]])
    return data

def teams_post(final_tickets):
    count = 0
    report_sections = []
    for user, items in final_tickets.items():
        for item in items:
            report_sections.extend([f'{user} [{item[0]}] {item[1]}'])
            count += 1
    myTeamsMessage = pymsteams.connectorcard(WEBHOOK_URL)
    ticket_section = pymsteams.cardsection()
    ticket_section.title(f'Aging tickets in queue: {count}')
    for content in report_sections:
        ticket_section.addFact(content,"")

    myTeamsMessage.addSection(ticket_section)
    # send the message.
    myTeamsMessage.text("Generated from Aging Tickets Auto Assignment Project")
    myTeamsMessage.send()

start_time = datetime.now()



if __name__ == "__main__":
    data_dict = {}
    with open("credentials.json") as reader:
        data_dict = json.loads(reader.read())
        snow_oauth_dict = data_dict["snow"]
    if DEBUG_MODE == True:
        snow_fqdn = "https://pgglobalenterpriseuat.service-now.com"
        snow_oauth_dict = snow_oauth_dict["snowNonProd"]
    else:
        snow_fqdn = "https://pgglobalenterprise.service-now.com"
        snow_oauth_dict = snow_oauth_dict["snowProd"]
    
    url = f"{snow_fqdn}/oauth_token.do"
    payload = (
        "refresh_token="
        + snow_oauth_dict["refreshToken"]
        + "&grant_type="
        + snow_oauth_dict["grant_type"]
        + "&client_id="
        + snow_oauth_dict["clientID"]
        + "&client_secret="
        + snow_oauth_dict["clientSecret"]
    )
    accessheaders = {
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.request("POST", url, headers=accessheaders, data=payload, verify=False).json()
    snowAccessToken = f'Bearer {response["access_token"]}'

    check_shift()

    time_6am = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
    time_1pm = current_time.replace(hour=13, minute=0, second=0, microsecond=0)
    time_10pm = current_time.replace(hour=22, minute=0, second=0, microsecond=0)
    engineer_on_shift = []
    time_window = timedelta(minutes=30)
    if time_6am - time_window <= current_time <= time_6am + time_window:     
        if DEBUG_MODE == True:
            getNewTicketUrl = "https://pgglobalenterpriseuat.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPART6%20am%20hour@javascript:gs.datePart(%27hour%27%2C%276%27%2C%27GE%27)%5Esys_created_onDATEPART5%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2717%27%2C%27LT%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"
        else:
            getNewTicketUrl = "https://pgglobalenterprise.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPART6%20am%20hour@javascript:gs.datePart(%27hour%27%2C%276%27%2C%27GE%27)%5Esys_created_onDATEPART5%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2717%27%2C%27LT%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"

        filterPayload={}
        filterHeaders = {
        'Authorization': snowAccessToken
        }

        response = requests.request("GET", getNewTicketUrl, headers=filterHeaders, data=filterPayload, verify=False).json()
        tickets = response
        shift_ticket = parse_tickets(tickets)
        engineer_on_shift = ap_on_shift

    elif time_1pm - time_window <= current_time <= time_1pm + time_window:
        if DEBUG_MODE == True:
            getNewTicketUrl = "https://pgglobalenterpriseuat.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPART5%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2717%27%2C%27GE%27)%5Esys_created_onDATEPART11%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2723%27%2C%27LE%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"
        else:
            getNewTicketUrl = "https://pgglobalenterprise.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPART5%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2717%27%2C%27GE%27)%5Esys_created_onDATEPART11%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2723%27%2C%27LE%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"

        filterPayload={}
        filterHeaders = {
        'Authorization': snowAccessToken
        }

        response = requests.request("GET", getNewTicketUrl, headers=filterHeaders, data=filterPayload, verify=False).json()
        tickets = response
        shift_ticket = parse_tickets(tickets)
        engineer_on_shift = eu_on_shift

    elif time_10pm - time_window <= current_time <= time_10pm + time_window:
        if DEBUG_MODE == True:
            getNewTicketUrl = "https://pgglobalenterpriseuat.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPARTNoon%20hour@javascript:gs.datePart(%27hour%27%2C%2712%27%2C%27GE%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Esys_created_onDATEPART6%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2718%27%2C%27LT%27)%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"
        else:
            getNewTicketUrl = "https://pgglobalenterprise.service-now.com/api/now/table/incident?sysparm_query=stateNOT%20IN6%2C7%5Esys_created_onNOTONLast%2030%20days@javascript:gs.beginningOfLast30Days()@javascript:gs.endOfLast30Days()%5Esys_created_onDATEPARTNoon%20hour@javascript:gs.datePart(%27hour%27%2C%2712%27%2C%27GE%27)%5Eassignment_group!%3Dd9348e92db8c5b081b5efe18bf9619d5%5Esys_created_onDATEPART6%20pm%20hour@javascript:gs.datePart(%27hour%27%2C%2718%27%2C%27LT%27)%5Eassignment_group%3Dfe908e1edbc85b081b5efe18bf961909%5EORassignment_group%3Dd6230e96db4c5b081b5efe18bf96195e%5Eassigned_toISEMPTY%5EORassigned_to%3D03a83d1fdb4017cc97985205dc961979&sysparm_first_row=1&sysparm_view=&sysparm_fields=number%2C%20short_description"

        filterPayload={}
        filterHeaders = {
        'Authorization': snowAccessToken
        }

        response = requests.request("GET", getNewTicketUrl, headers=filterHeaders, data=filterPayload, verify=False).json()
        tickets = response
        shift_ticket = parse_tickets(tickets)
        engineer_on_shift = us_on_shift
    if engineer_on_shift == []:
        print("You are running this code outside designated hours. Make sure to run the code at the start of every shift.")
        exit()
    final_assigned = assign_ticket(engineer_on_shift,shift_ticket)
    teams_post(final_assigned)
    print("Total Elapsed time: " + str(datetime.now() - start_time))

    